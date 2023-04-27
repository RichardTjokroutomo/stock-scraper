import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time

EXCEL_FILENAME = "data.xlsx"
EXCEL_PATH = "YOUR_PATH _HERE"
#EXCEL_PATH = "C:\\Users\\Richard\\Documents\\-projects\\project-15\\"
ROOT_URL = "https://finance.yahoo.com/quote/"
DRIVER_PATH = "[YOUR_PATH_HERE]/chromedriver.exe"
#DRIVER_PATH = "C:\Program Files (x86)\chromedriver.exe"

# get the user data
# ====================================================================================================
wb_obj = openpyxl.load_workbook(EXCEL_PATH + EXCEL_FILENAME)

sheet_obj = wb_obj.active

ticker_code = sheet_obj.cell(row = 2, column = 1)




# get the stock price
# ====================================================================================================
driver = webdriver.Chrome(DRIVER_PATH)
URL = ROOT_URL + ticker_code.value;

print(URL)

driver.get(URL)
print(driver.title)

# fill the name
name_xpath = '//*[@id="quote-header-info"]/div[3]/div[1]/div[1]/fin-streamer[1]'
price_handler = driver.find_element(By.XPATH, name_xpath)
price = price_handler.get_attribute("value")
print("price of the stock: " + price)
#time.sleep(1)


driver.quit()


# update the price on excel
# ====================================================================================================
print("now updating the value")

ws = wb_obj['Sheet1']
price_cell = ws['B2']
price_cell.value = price
wb_obj.save(EXCEL_FILENAME)